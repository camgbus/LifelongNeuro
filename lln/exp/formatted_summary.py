import pandas as pd
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle

def shape_format_summarized_df(path, orig_file_name, new_file_name, datasets, metrics, higher_is_better, exp_better_names=None):
    summarized_results, first_best_metric_means, second_best_metric_means = shape_summarized_df(path, orig_file_name, datasets, metrics, higher_is_better, exp_better_names=exp_better_names)
    first_best_indexes, second_best_indexes = get_best_indexes(summarized_results, datasets, metrics, first_best_metric_means, second_best_metric_means)
    format_summarized_df(summarized_results, first_best_indexes, second_best_indexes, path, new_file_name)

def shape_summarized_df(path, orig_file_name, datasets, metrics, higher_is_better, exp_better_names=None):
    summarized_results = []
    new_columns = ["Exp"] + [f"{d} {m}" for d, m in [(d, m) for d in datasets for m in metrics]]
    file = os.path.join(path, orig_file_name+".csv")
    df = pd.read_csv(file)
    df = df[df['Dataset'].isin(datasets)]
    exp_names = sorted(list(df['Exp'].unique()))
    first_best_metric_means = {ds: {m: (0, None)  if higher_is_better[m] else (1000, None) for m in metrics} for ds in datasets}
    second_best_metric_means = {ds: {m: (0, None)  if higher_is_better[m] else (1000, None) for m in metrics} for ds in datasets}
    if exp_better_names is None:
        exp_better_names = {x: x for x in exp_names}
    for exp in exp_names:
        exp_row = [exp_better_names[exp]]
        for ds in datasets:
            mean_row = df[(df['Exp'] == exp) & (df['Dataset'] == ds) & (df['Run'] == 'Mean')]
            std_row = df[(df['Exp'] == exp) & (df['Dataset'] == ds) & (df['Run'] == 'Std')]
            assert len(mean_row) == len(std_row) == 1
            mean_row, std_row = mean_row.iloc[0], std_row.iloc[0]    
            exp_row += [f"{mean_row[m]:.2f} ± {std_row[m]:.2f}" for m in metrics]
            # Formatting
            for m in metrics:
                if higher_is_better[m]:
                    if mean_row[m] > first_best_metric_means[ds][m][0]:
                        second_best_metric_means[ds][m] = first_best_metric_means[ds][m]
                        first_best_metric_means[ds][m] = (mean_row[m], exp)
                    elif mean_row[m] > second_best_metric_means[ds][m][0]:
                        second_best_metric_means[ds][m] = (mean_row[m], exp)
                else:
                    if mean_row[m] < first_best_metric_means[ds][m][0]:
                        second_best_metric_means[ds][m] = first_best_metric_means[ds][m]
                        first_best_metric_means[ds][m] = (mean_row[m], exp)
                    elif mean_row[m] < second_best_metric_means[ds][m][0]:
                        second_best_metric_means[ds][m] = (mean_row[m], exp)
        summarized_results.append(exp_row)
    summarized_results = pd.DataFrame(summarized_results, columns=new_columns)
    return summarized_results, first_best_metric_means, second_best_metric_means

def get_best_indexes(summarized_results, datasets, metrics, first_best_metric_means, second_best_metric_means):
    # Find the indexes of cells that must be highlighted as best
    first_best_indexes = []
    second_best_indexes = []
    for ds in datasets:
        for m in metrics:
            first_best_exp = first_best_metric_means[ds][m][1]
            second_best_exp = second_best_metric_means[ds][m][1]
            if first_best_exp is not None:
                row_index = summarized_results.loc[summarized_results['Exp'] == first_best_exp].index[0]
                column_index = summarized_results.columns.get_loc(f"{ds} {m}")
                first_best_indexes.append((row_index+2, column_index+1))
            if second_best_exp is not None:
                row_index = summarized_results.loc[summarized_results['Exp'] == second_best_exp].index[0]
                column_index = summarized_results.columns.get_loc(f"{ds} {m}")
                second_best_indexes.append((row_index+2, column_index+1))  
    return first_best_indexes, second_best_indexes  

def format_summarized_df(df, first_best_indexes, second_best_indexes, path, new_file_name):
    # Set up the Excel writer
    writer = pd.ExcelWriter(os.path.join(path, f'{new_file_name}.xlsx'), engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    
    # Access the workbook and the worksheet
    workbook = writer.book
    if workbook is None:
        workbook = Workbook()
        writer.book = workbook

    worksheet = writer.sheets['Sheet1']
    
    # Define a NamedStyle for highlighting
    first_best = NamedStyle(name="first_best", font=Font(bold=True, italic=True, color="36A556"))
    second_best = NamedStyle(name="second_best", font=Font(bold=True, color="5D69C3"))

    # Add the NamedStyle to the workbook to use it
    workbook.add_named_style(first_best)
    workbook.add_named_style(second_best)
    
    for row_index, column_index in first_best_indexes:
        worksheet.cell(row=row_index, column=column_index).style = first_best
    for row_index, column_index in second_best_indexes:
        worksheet.cell(row=row_index, column=column_index).style = second_best
    
    # Save and close the workbook
    writer.close()